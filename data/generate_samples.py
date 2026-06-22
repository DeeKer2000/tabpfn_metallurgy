"""
合成炉料样本生成器
用途：为铜熔炼热力学计算生成批量合成炉料成分数据
"""

import pandas as pd
import numpy as np
import openpyxl

# ============================================================
#                        参 数 设 置
# ============================================================

# 矿源文件路径
ORE_FILE = "矿源.xlsx"

# 生成样本数量
N_SAMPLES = 30000

# 随机种子（设为None则每次生成不同数据）
RANDOM_SEED = 42

# 每组选用的矿种数量范围
N_ORE_MIN = 2
N_ORE_MAX = 6

# SiO2添加量范围（占矿石质量的百分比）
SIO2_ADD_MIN = 0.0    # %
SIO2_ADD_MAX = 15.0   # %

# ---- 温度设置 ----
# 设为None则在TEMP_MIN~TEMP_MAX之间随机
# 设为具体值（如1300.0）则所有样本使用同一温度
TEMP_FIXED = None
TEMP_MIN = 1190.0
TEMP_MAX = 1350.0

# ---- 压强设置 ----
# 固定值，单位atm（铜熔炼通常为常压1 atm）
PRESSURE = 1.0

# 冶金成分约束（加入SiO2后的最终成分）
CONSTRAINTS = {
    "Cu":      (14.0, 26.0),
    "S":       (8.0,  38.0),
    "Fe":      (8.0,  36.0),
    "SiO2":    (3.0,  25.0),
    "Fe/SiO2": (0.8,  4.0),
}

# 输出文件
OUTPUT_EXCEL = "合成炉料样本.xlsx"
OUTPUT_TXT = "合成炉料样本.txt"
SHEET_NAME = "FurnaceCharge"

# ============================================================
#                        生成逻辑
# ============================================================

COMP_NAMES = ['Cu', 'S', 'Fe', 'SiO2', 'Pb', 'Zn', 'MgO', 'Al2O3', 'Ni', 'Sb', 'Bi', 'CaO', 'As']
COMP_COL_IDX = [2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14]

SIO2_IDX = COMP_NAMES.index('SiO2')
FE_IDX = COMP_NAMES.index('Fe')
CU_IDX = COMP_NAMES.index('Cu')
S_IDX = COMP_NAMES.index('S')


def load_ore_data(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    data_rows = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        if row[0] is not None:
            data_rows.append(list(row))
    wb.close()
    ore_names = [r[0] for r in data_rows]
    ore_comp = np.array([[float(r[i]) for i in COMP_COL_IDX] for r in data_rows])
    return ore_names, ore_comp


def check_constraints(final_comp, final_sio2):
    cu = final_comp[CU_IDX]
    s = final_comp[S_IDX]
    fe = final_comp[FE_IDX]
    fe_sio2 = fe / final_sio2 if final_sio2 > 0 else 999

    lo, hi = CONSTRAINTS['Cu']
    if not (lo <= cu <= hi): return False
    lo, hi = CONSTRAINTS['S']
    if not (lo <= s <= hi): return False
    lo, hi = CONSTRAINTS['Fe']
    if not (lo <= fe <= hi): return False
    lo, hi = CONSTRAINTS['SiO2']
    if not (lo <= final_sio2 <= hi): return False
    lo, hi = CONSTRAINTS['Fe/SiO2']
    if not (lo <= fe_sio2 <= hi): return False
    return True


def generate_samples(ore_names, ore_comp, n_samples, seed=None):
    if seed is not None:
        np.random.seed(seed)

    n_ores = len(ore_names)
    records = []
    attempts = 0
    max_attempts = n_samples * 200

    while len(records) < n_samples and attempts < max_attempts:
        attempts += 1

        n_select = np.random.randint(N_ORE_MIN, N_ORE_MAX + 1)
        selected_idx = np.random.choice(n_ores, size=n_select, replace=False)
        selected_comp = ore_comp[selected_idx]
        weights = np.random.dirichlet(np.ones(n_select))

        mixed = np.zeros(len(COMP_NAMES))
        for k in range(n_select):
            mixed += weights[k] * selected_comp[k]

        ore_sio2 = mixed[SIO2_IDX]
        add_sio2 = np.random.uniform(SIO2_ADD_MIN, SIO2_ADD_MAX) / 100.0
        total_mass = 1.0 + add_sio2

        final_comp = mixed.copy()
        for j in range(len(COMP_NAMES)):
            if j == SIO2_IDX:
                final_comp[j] = (mixed[j] + add_sio2 * 100.0) / total_mass
            else:
                final_comp[j] = mixed[j] / total_mass

        final_sio2 = final_comp[SIO2_IDX]

        if not check_constraints(final_comp, final_sio2):
            continue

        # 温度
        if TEMP_FIXED is not None:
            temp = TEMP_FIXED
        else:
            temp = np.random.uniform(TEMP_MIN, TEMP_MAX)

        # 组装记录：温度、压强放最前面
        rec = {
            'No': len(records) + 1,
            '温度_℃': round(temp, 1),
            '压强_atm': PRESSURE,
        }
        for j in range(n_ores):
            if j in selected_idx:
                pos = list(selected_idx).index(j)
                rec[ore_names[j]] = round(weights[pos] * 100, 2)
            else:
                rec[ore_names[j]] = 0.0
        rec['SiO2_原矿'] = round(ore_sio2, 3)
        rec['SiO2_添加量'] = round(add_sio2 * 100, 3)
        for j, c in enumerate(COMP_NAMES):
            rec[c] = round(final_comp[j], 3)
        records.append(rec)

    return records, attempts


def main():
    print(f"矿源文件:   {ORE_FILE}")
    print(f"生成数量:   {N_SAMPLES}")
    if TEMP_FIXED is not None:
        print(f"温度:       固定 {TEMP_FIXED} ℃")
    else:
        print(f"温度区间:   {TEMP_MIN}~{TEMP_MAX} ℃（随机）")
    print(f"压强:       {PRESSURE} atm")
    print(f"SiO2添加:   {SIO2_ADD_MIN}~{SIO2_ADD_MAX} %")
    print()

    ore_names, ore_comp = load_ore_data(ORE_FILE)
    print(f"读取到 {len(ore_names)} 种矿: {ore_names}")
    print()

    records, attempts = generate_samples(ore_names, ore_comp, N_SAMPLES, RANDOM_SEED)

    out = pd.DataFrame(records)

    # 输出Excel
    out.to_excel(OUTPUT_EXCEL, index=False, sheet_name=SHEET_NAME)

    # 输出TXT（Tab分隔，Excel可直接打开）
    out.to_csv(OUTPUT_TXT, index=False, sep='\t', encoding='utf-8-sig')

    ore_sum = out[ore_names].sum(axis=1)
    print(f"生成完成: {len(records)} 组, 尝试 {attempts} 次")
    print(f"矿石配比总和: {ore_sum.min():.2f} ~ {ore_sum.max():.2f}")
    print(f"SiO2原矿:     {out['SiO2_原矿'].min():.2f} ~ {out['SiO2_原矿'].max():.2f}")
    print(f"SiO2添加量:   {out['SiO2_添加量'].min():.2f} ~ {out['SiO2_添加量'].max():.2f}")
    print(f"SiO2最终:     {out['SiO2'].min():.2f} ~ {out['SiO2'].max():.2f}")
    print(f"温度:         {out['温度_℃'].min():.1f} ~ {out['温度_℃'].max():.1f} ℃")
    print(f"压强:         {PRESSURE} atm")
    print()
    print(f"已保存: {OUTPUT_EXCEL}")
    print(f"已保存: {OUTPUT_TXT}")


if __name__ == '__main__':
    main()
